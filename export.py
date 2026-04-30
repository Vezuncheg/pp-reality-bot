"""export.py — выгрузка БД в Excel (3 листа)"""
import io
from db import get_all_payments, get_all_users, get_all_events


def export_to_excel() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Стили ──────────────────────────────────────────────
    def make_header(ws, headers, color="7C3AED"):
        fill = PatternFill("solid", fgColor=color)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    def set_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # ── Лист 1: Платежи ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Платежи"
    headers1 = ["#", "Имя", "Email", "Telegram ID", "Username",
                "Тариф", "Сумма ₽", "Дата оплаты", "Клуб до", "Статус"]
    make_header(ws1, headers1)
    for i, row in enumerate(get_all_payments(), 2):
        for j, val in enumerate(row, 1):
            ws1.cell(row=i, column=j, value=val)
    set_widths(ws1, [5, 20, 25, 15, 15, 20, 10, 20, 20, 10])

    # ── Лист 2: Пользователи ───────────────────────────────
    ws2 = wb.create_sheet("Пользователи")
    headers2 = ["Telegram ID", "Username", "Имя", "Первый визит", "Последний визит",
                "Запусков", "Архетип", "Пол", "Возраст", "Вес", "Рост", "Цель",
                "Тариф", "Сумма ₽", "Дата оплаты"]
    make_header(ws2, headers2, color="059669")

    arch_names = {
        "emotional_eater": "Эмоц. едок",
        "social_hostage": "Соц. заложник",
        "metabolic_skeptic": "Метаб. скептик",
        "starter_stopper": "Стартер-стопер",
    }
    goal_names = {
        "fat": "Убрать жир",
        "muscle": "Набрать мышцы",
        "tone": "Рельеф/тонус",
    }
    gender_names = {"male": "Мужчина", "female": "Женщина"}

    for i, row in enumerate(get_all_users(), 2):
        (tg_id, username, full_name, first_seen, last_seen,
         start_count, archetype, gender, age, weight, height, goal,
         plan, amount, paid_at) = row

        ws2.cell(row=i, column=1, value=tg_id)
        ws2.cell(row=i, column=2, value=f"@{username}" if username else "—")
        ws2.cell(row=i, column=3, value=full_name or "—")
        ws2.cell(row=i, column=4, value=str(first_seen)[:16] if first_seen else "—")
        ws2.cell(row=i, column=5, value=str(last_seen)[:16] if last_seen else "—")
        ws2.cell(row=i, column=6, value=start_count)
        ws2.cell(row=i, column=7, value=arch_names.get(archetype, archetype or "—"))
        ws2.cell(row=i, column=8, value=gender_names.get(gender, gender or "—"))
        ws2.cell(row=i, column=9, value=age or "—")
        ws2.cell(row=i, column=10, value=float(weight) if weight else "—")
        ws2.cell(row=i, column=11, value=float(height) if height else "—")
        ws2.cell(row=i, column=12, value=goal_names.get(goal, goal or "—"))
        ws2.cell(row=i, column=13, value=plan)
        ws2.cell(row=i, column=14, value=amount)
        ws2.cell(row=i, column=15, value=str(paid_at)[:16] if paid_at and paid_at != "—" else "—")

    set_widths(ws2, [15, 18, 20, 18, 18, 8, 18, 10, 8, 8, 8, 15, 20, 10, 18])

    # ── Лист 3: События ────────────────────────────────────
    ws3 = wb.create_sheet("События")
    headers3 = ["Telegram ID", "Имя", "Username", "Событие", "Данные", "Время"]
    make_header(ws3, headers3, color="DC2626")

    event_labels = {
        "start": "🚀 Запуск бота",
        "quiz_completed": "✅ Анкета завершена",
        "block_viewed": "👁 Блок просмотрен",
        "pay_clicked": "💳 Клик на оплату",
        "paid": "💰 Оплата",
        "broadcast_sent": "📢 Рассылка",
        "support_message": "💬 Сообщение в поддержку",
    }

    for i, row in enumerate(get_all_events(), 2):
        tg_id, full_name, username, event, data, created_at = row
        ws3.cell(row=i, column=1, value=tg_id)
        ws3.cell(row=i, column=2, value=full_name or "—")
        ws3.cell(row=i, column=3, value=f"@{username}" if username else "—")
        ws3.cell(row=i, column=4, value=event_labels.get(event, event))
        ws3.cell(row=i, column=5, value=data or "—")
        ws3.cell(row=i, column=6, value=str(created_at)[:16] if created_at else "—")

    set_widths(ws3, [15, 20, 18, 22, 20, 18])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
